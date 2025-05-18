import os
import sys
import json
from pathlib import Path
from dotenv import load_dotenv
import gspread
from oauth2client.service_account import ServiceAccountCredentials
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, Border, Side
from openpyxl.utils import get_column_letter
from collections import defaultdict
from datetime import datetime
from PyQt5.QtWidgets import (QApplication, QMainWindow, QVBoxLayout, QWidget, QPushButton,
                             QLabel, QProgressBar, QFileDialog, QMessageBox, QTextEdit,
                             QHBoxLayout, QLineEdit, QComboBox)
from PyQt5.QtCore import Qt, QThread, pyqtSignal
from PyQt5.QtGui import QPalette, QColor


class GoogleSheetsWorker(QThread):
    """Поток для обработки данных из Google Sheets"""
    progress = pyqtSignal(int)
    message = pyqtSignal(str)
    finished = pyqtSignal(object, object, object)
    error = pyqtSignal(str)

    def __init__(self, spreadsheet_url, credentials_file, sheet_name):
        super().__init__()
        self.spreadsheet_url = spreadsheet_url
        self.credentials_file = credentials_file
        self.sheet_name = sheet_name

    def run(self):
        try:
            self.message.emit("Получение данных из Google Sheets...")
            raw_data = self.get_google_sheets_data()
            self.progress.emit(30)

            self.message.emit("Обработка данных...")
            address_data, camera_models, object_codes = self.process_camera_data(
                raw_data)
            self.progress.emit(70)

            self.finished.emit(address_data, camera_models, object_codes)
            self.progress.emit(100)
            self.message.emit("Данные успешно обработаны!")
        except Exception as e:
            self.error.emit(f"Ошибка: {str(e)}")

    def get_google_sheets_data(self):
        """Получение данных из Google Sheets"""
        scope = ["https://spreadsheets.google.com/feeds",
                 "https://www.googleapis.com/auth/drive"]

        if not os.path.exists(self.credentials_file):
            raise FileNotFoundError(
                f"Файл ключей {self.credentials_file} не найден!")

        credentials = ServiceAccountCredentials.from_json_keyfile_name(
            self.credentials_file, scope)
        client = gspread.authorize(credentials)
        spreadsheet = client.open_by_url(self.spreadsheet_url)
        worksheet = spreadsheet.worksheet(self.sheet_name)

        expected_headers = ["Код объекта", "Адрес установки", "Камера"]
        return worksheet.get_all_records(expected_headers=expected_headers)

    def process_camera_data(self, data):
        """Обработка и группировка данных по адресам"""
        if not data:
            raise ValueError("В таблице нет данных!")

        address_data = defaultdict(lambda: defaultdict(int))
        all_models = set()
        object_codes = {}

        for row in data:
            code = row.get("Код объекта", "").strip()
            address = row.get("Адрес установки", "").strip()
            model = row.get("Камера", "").strip()
            if address and model:
                address_data[address][model] += 1
                all_models.add(model)
                object_codes[address] = code

        if not address_data:
            raise ValueError("Нет данных для формирования отчета!")

        return address_data, sorted(all_models), object_codes


class ExcelReportGenerator(QThread):
    """Поток для генерации Excel отчета"""
    progress = pyqtSignal(int)
    message = pyqtSignal(str)
    finished = pyqtSignal(str)
    error = pyqtSignal(str)

    def __init__(self, address_data, camera_models, object_codes):
        super().__init__()
        self.address_data = address_data
        self.camera_models = camera_models
        self.object_codes = object_codes

    def run(self):
        try:
            self.message.emit("Создание Excel отчета...")
            report = self.create_excel_report()
            self.progress.emit(50)

            filename = self.save_report(report)
            self.progress.emit(100)
            self.finished.emit(filename)
            self.message.emit("Отчет успешно создан!")
        except Exception as e:
            self.error.emit(f"Ошибка при создании отчета: {str(e)}")

    def create_excel_report(self):
        """Создание Excel файла с отчетом"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Лист1"

        # Стили оформления
        header_style = Font(bold=True)
        horizontal_alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True)
        vertical_alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True, text_rotation=90)
        border_style = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        wrap_alignment = Alignment(wrap_text=True)

        # Основной заголовок (A1:U1)
        ws.merge_cells('A1:U1')
        ws['A1'] = "Ведомость установленного и замонтированного оборудования по объекту:"
        ws['A1'].font = header_style
        ws['A1'].alignment = horizontal_alignment

        # Название проекта (A2:U2)
        ws.merge_cells('A2:U2')
        ws['A2'] = "Реконструкция местных линий связи к объектам РСМОБ г. Бреста перекрестки, 8 этап"
        ws['A2'].font = header_style
        ws['A2'].alignment = horizontal_alignment

        # Группы оборудования (строка 3)
        ws.merge_cells('C3:M3')
        ws['C3'] = "Видеокамеры"
        ws['C3'].alignment = horizontal_alignment

        ws.merge_cells('Q3:S3')
        ws['Q3'] = "Коммутаторы"
        ws['Q3'].alignment = horizontal_alignment

        ws.merge_cells('T3:U3')
        ws['T3'] = "Удлинитель"
        ws['T3'].alignment = horizontal_alignment

        # Заголовки столбцов (строка 4)
        headers = [
            "Код объекта",
            "АДРЕС",
            "(2 Мп) TIANDY TC-C32GS-I5EYCSD (2.8mm/V4.2)",
            "(2МР) IPC2122LB-ADF28KM-G",
            "DS-2CD1043G0-IUVSD 4mm",
            "DS-2CD2123G2-IUVSD 4mm",
            "DS-2CD2T23G2-2IUVSD",
            "DS-2CD2T23G2-2IUVSD 4mm",
            "DS-2CD3021G0-IUVSC 4mm",
            "DS-2CD3123G2-IUUVSC 6mm",
            "DS-2CD3626G2T-IZSUVSC (7-35mm)",
            "DS-2CD3626G2T-IZSUVSC 7-35mm",
            "DS-2CD3726G2T-IZSUVSC (7-35mm)",
            "DS-2DE5425IW-AEUVSC",
            "HIKVISION DS-2DE5425IW-A E (T5)",
            "Uniview IPC2122LE-ADF28KMC-WL",
            "Коммутатор ZTO L2S1900-4TP2S",
            "Коммутатор ZTO L2S 1900-8TP2S",
            "Коммутатор ZTO L2S 1900-16TP2S",
            "Инжектор питания (PoE) OPL-POE-Ex802 3at-100-IP67",
            "Удлинитель PoE ZTO POEEXT 100"
        ]

        # Заполняем строку 4 с разным выравниванием
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col, value=header)
            cell.font = header_style
            cell.border = border_style
            cell.alignment = vertical_alignment if col >= 3 else horizontal_alignment

        ws.row_dimensions[4].height = 150

        # Заполнение данных
        row_idx = 5
        for address, counts in self.address_data.items():
            ws.cell(row=row_idx, column=1,
                    value=self.object_codes.get(address, ""))
            ws.cell(row=row_idx, column=2, value=address)

            for model, quantity in counts.items():
                if model in headers:
                    col_idx = headers.index(model) + 1
                    ws.cell(row=row_idx, column=col_idx, value=quantity)

            for col in range(17, 22):
                ws.cell(row=row_idx, column=col, value="")

            row_idx += 1

        # Итоговая строка
        total_row = row_idx
        ws.cell(row=total_row, column=1, value="ИТОГО:")
        ws.cell(row=total_row, column=2, value="")

        for col in range(3, 22):
            col_letter = get_column_letter(col)
            ws.cell(row=total_row, column=col,
                    value=f"=SUM({col_letter}5:{col_letter}{total_row-1})")

        # Подпись
        signature_row = total_row + 1
        ws.merge_cells(f'A{signature_row}:U{signature_row}')
        ws[f'A{signature_row}'] = "Подготовил: ведущий инженер ЛСС и АУ А.И. Козей"
        ws[f'A{signature_row}'].alignment = horizontal_alignment

        # Форматирование ячеек
        for row in ws.iter_rows(min_row=1, max_row=signature_row, min_col=1, max_col=21):
            for cell in row:
                cell.alignment = wrap_alignment
                cell.border = border_style

        # Ширина столбцов
        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 50
        for col in range(3, 22):
            ws.column_dimensions[get_column_letter(col)].width = 5

        return wb

    def save_report(self, report):
        """Сохранение отчета в файл"""
        output_dir = Path("output")
        output_dir.mkdir(exist_ok=True)

        first_code = next(iter(self.object_codes.values()),
                          "") if self.object_codes else ""
        id_number = first_code[1] if len(
            first_code) > 1 and first_code[1].isdigit() else ""

        filename = f"Ведомость-{id_number}.xlsx" if id_number else f"Ведомость-{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        filepath = output_dir / filename
        report.save(filepath)
        return str(filepath)


class MainWindow(QMainWindow):
    """Главное окно приложения"""

    def __init__(self):
        super().__init__()
        self.setWindowTitle("Генератор ведомостей оборудования")
        self.setMinimumSize(800, 600)

        # Загрузка конфигурации
        load_dotenv()
        self.spreadsheet_url = os.getenv("GOOGLE_SHEETS_URL", "")
        self.credentials_file = os.getenv(
            "CREDENTIALS_JSON", "credentials.json")
        self.sheet_name = os.getenv("SHEET_NAME", "Камеры")
        self.client_email = ""

        self.init_ui()
        self.load_credentials_info()
        self.apply_theme("light")  # По умолчанию светлая тема

    def init_ui(self):
        """Инициализация интерфейса"""
        central_widget = QWidget()
        self.setCentralWidget(central_widget)

        layout = QVBoxLayout()
        central_widget.setLayout(layout)

        # Выбор темы
        theme_layout = QHBoxLayout()
        theme_label = QLabel("Тема:")
        self.theme_combo = QComboBox()
        self.theme_combo.addItems(["light", "dark"])
        self.theme_combo.currentTextChanged.connect(self.apply_theme)
        theme_layout.addWidget(theme_label)
        theme_layout.addWidget(self.theme_combo)
        layout.addLayout(theme_layout)

        # Поля для настроек
        settings_layout = QVBoxLayout()

        # URL Google Sheets
        url_layout = QHBoxLayout()
        url_label = QLabel("URL Google Sheets:")
        self.url_edit = QLineEdit(self.spreadsheet_url)
        self.url_edit.textChanged.connect(self.update_spreadsheet_url)
        url_layout.addWidget(url_label)
        url_layout.addWidget(self.url_edit)
        settings_layout.addLayout(url_layout)

        # Имя листа
        sheet_layout = QHBoxLayout()
        sheet_label = QLabel("Имя листа:")
        self.sheet_edit = QLineEdit(self.sheet_name)
        self.sheet_edit.textChanged.connect(self.update_sheet_name)
        sheet_layout.addWidget(sheet_label)
        sheet_layout.addWidget(self.sheet_edit)
        settings_layout.addLayout(sheet_layout)

        # Файл учетных данных
        creds_layout = QHBoxLayout()
        creds_label = QLabel("Файл учетных данных:")
        self.creds_edit = QLineEdit(self.credentials_file)
        self.creds_edit.textChanged.connect(self.update_credentials_file)
        creds_button = QPushButton("Обзор...")
        creds_button.clicked.connect(self.browse_credentials_file)
        creds_layout.addWidget(creds_label)
        creds_layout.addWidget(self.creds_edit)
        creds_layout.addWidget(creds_button)
        settings_layout.addLayout(creds_layout)

        # Client email (только для чтения)
        email_layout = QHBoxLayout()
        email_label = QLabel("Client Email:")
        self.email_display = QLineEdit()
        self.email_display.setReadOnly(True)
        self.email_display.setStyleSheet("background-color: #f0f0f0;")
        copy_button = QPushButton("Копировать")
        copy_button.clicked.connect(self.copy_client_email)
        email_layout.addWidget(email_label)
        email_layout.addWidget(self.email_display)
        email_layout.addWidget(copy_button)
        settings_layout.addLayout(email_layout)

        layout.addLayout(settings_layout)

        # Лог сообщений
        self.log = QTextEdit()
        self.log.setReadOnly(True)
        layout.addWidget(self.log)

        # Прогресс бар
        self.progress = QProgressBar()
        layout.addWidget(self.progress)

        # Кнопка запуска
        self.run_btn = QPushButton("Создать ведомость")
        self.run_btn.clicked.connect(self.run_report_generation)
        layout.addWidget(self.run_btn)

        # Статус
        self.status_label = QLabel()
        layout.addWidget(self.status_label)

    def load_credentials_info(self):
        """Загрузка информации из файла учетных данных"""
        try:
            if os.path.exists(self.credentials_file):
                with open(self.credentials_file, 'r') as f:
                    creds = json.load(f)
                    self.client_email = creds.get('client_email', '')
                    self.email_display.setText(self.client_email)
        except Exception as e:
            self.log_message(f"Ошибка загрузки файла учетных данных: {str(e)}")

    def apply_theme(self, theme_name):
        """Применение выбранной темы"""
        palette = QPalette()

        if theme_name == "dark":
            palette.setColor(QPalette.Window, QColor(53, 53, 53))
            palette.setColor(QPalette.WindowText, Qt.white)
            palette.setColor(QPalette.Base, QColor(25, 25, 25))
            palette.setColor(QPalette.AlternateBase, QColor(53, 53, 53))
            palette.setColor(QPalette.ToolTipBase, Qt.white)
            palette.setColor(QPalette.ToolTipText, Qt.white)
            palette.setColor(QPalette.Text, Qt.white)
            palette.setColor(QPalette.Button, QColor(53, 53, 53))
            palette.setColor(QPalette.ButtonText, Qt.white)
            palette.setColor(QPalette.BrightText, Qt.red)
            palette.setColor(QPalette.Link, QColor(42, 130, 218))
            palette.setColor(QPalette.Highlight, QColor(42, 130, 218))
            palette.setColor(QPalette.HighlightedText, Qt.black)
        else:
            palette = QApplication.style().standardPalette()

        QApplication.setPalette(palette)

    def update_spreadsheet_url(self, url):
        """Обновление URL таблицы"""
        self.spreadsheet_url = url.strip()

    def update_sheet_name(self, name):
        """Обновление имени листа"""
        self.sheet_name = name.strip()

    def update_credentials_file(self, path):
        """Обновление пути к файлу учетных данных"""
        self.credentials_file = path.strip()
        self.load_credentials_info()

    def browse_credentials_file(self):
        """Выбор файла учетных данных через диалог"""
        filename, _ = QFileDialog.getOpenFileName(
            self, "Выберите файл учетных данных", "", "JSON Files (*.json)")
        if filename:
            self.creds_edit.setText(filename)
            self.credentials_file = filename
            self.load_credentials_info()

    def copy_client_email(self):
        """Копирование client_email в буфер обмена"""
        clipboard = QApplication.clipboard()
        clipboard.setText(self.client_email)
        self.log_message("Client email скопирован в буфер обмена")

    def log_message(self, message):
        """Добавление сообщения в лог"""
        self.log.append(message)
        self.log.ensureCursorVisible()

    def run_report_generation(self):
        """Запуск процесса генерации отчета"""
        if not all([self.spreadsheet_url, self.credentials_file]):
            self.log_message("Ошибка: Не заданы все необходимые параметры!")
            return

        self.log_message("Начало обработки данных...")
        self.progress.setValue(0)
        self.run_btn.setEnabled(False)

        # Создаем и запускаем worker для получения данных
        self.sheets_worker = GoogleSheetsWorker(
            self.spreadsheet_url,
            self.credentials_file,
            self.sheet_name
        )

        # Подключаем сигналы
        self.sheets_worker.progress.connect(self.progress.setValue)
        self.sheets_worker.message.connect(self.log_message)
        self.sheets_worker.finished.connect(self.on_data_processed)
        self.sheets_worker.error.connect(self.on_error)

        self.sheets_worker.start()

    def on_data_processed(self, address_data, camera_models, object_codes):
        """Обработка завершения получения данных"""
        self.log_message(f"Обработано {len(address_data)} адресов")
        self.log_message(f"Найдено {len(camera_models)} моделей камер")

        # Создаем и запускаем worker для генерации отчета
        self.report_worker = ExcelReportGenerator(
            address_data,
            camera_models,
            object_codes
        )

        # Подключаем сигналы
        self.report_worker.progress.connect(self.progress.setValue)
        self.report_worker.message.connect(self.log_message)
        self.report_worker.finished.connect(self.on_report_generated)
        self.report_worker.error.connect(self.on_error)

        self.report_worker.start()

    def on_report_generated(self, filename):
        """Обработка завершения генерации отчета"""
        self.run_btn.setEnabled(True)
        self.status_label.setText(f"Отчет сохранен: {filename}")

        # Показать сообщение об успехе
        msg = QMessageBox()
        msg.setIcon(QMessageBox.Information)
        msg.setText("Отчет успешно создан!")
        msg.setInformativeText(f"Файл сохранен как: {filename}")
        msg.setWindowTitle("Успех")
        msg.exec_()

    def on_error(self, error_message):
        """Обработка ошибок"""
        self.log_message(error_message)
        self.run_btn.setEnabled(True)
        self.progress.setValue(0)

        # Показать сообщение об ошибке
        msg = QMessageBox()
        msg.setIcon(QMessageBox.Critical)
        msg.setText("Произошла ошибка!")
        msg.setInformativeText(error_message)
        msg.setWindowTitle("Ошибка")
        msg.exec_()


def main():
    """Точка входа в приложение"""
    # Для корректного отображения в Windows
    if sys.platform == "win32":
        import ctypes
        ctypes.windll.shell32.SetCurrentProcessExplicitAppUserModelID(
            "EquipmentReportGenerator.1.0")

    app = QApplication(sys.argv)
    app.setStyle('Fusion')  # Установка стиля для кроссплатформенного вида

    # Создаем папку для отчетов, если ее нет
    output_dir = Path("output")
    output_dir.mkdir(exist_ok=True)

    window = MainWindow()
    window.show()

    sys.exit(app.exec_())


if __name__ == "__main__":
    main()
