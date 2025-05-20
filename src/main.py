import os
import sys
import json
from pathlib import Path
from dotenv import load_dotenv
import gspread
from oauth2client.service_account import ServiceAccountCredentials
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, Border, Side
from openpyxl.utils import get_column_letter
from collections import defaultdict
from datetime import datetime
import threading
import tkinter as tk
from tkinter import ttk, messagebox, filedialog, scrolledtext
from tkinter.font import Font as TkFont


class GoogleSheetsWorker:
    """Класс для обработки данных из Google Sheets"""

    def __init__(self, spreadsheet_url, credentials_file, sheet_name, callback):
        self.spreadsheet_url = spreadsheet_url
        self.credentials_file = credentials_file
        self.sheet_name = sheet_name
        self.callback = callback

    def run(self):
        try:
            raw_data = self.get_google_sheets_data()
            address_data, camera_models, object_codes = self.process_camera_data(
                raw_data)
            self.callback(address_data, camera_models, object_codes, None)
        except Exception as e:
            self.callback(None, None, None, str(e))

    def get_google_sheets_data(self):
        """Получение данных из Google Sheets"""
        scope = ["https://spreadsheets.google.com/feeds",
                 "https://www.googleapis.com/auth/drive"]

        if not os.path.exists(self.credentials_file):
            raise FileNotFoundError(
                f"Файл ключей {self.credentials_file} не найден!")

        credentials = ServiceAccountCredentials.from_json_keyfile_name(
            self.credentials_file, scope)
        client = gspread.authorize(credentials)
        spreadsheet = client.open_by_url(self.spreadsheet_url)
        worksheet = spreadsheet.worksheet(self.sheet_name)

        expected_headers = ["Код объекта", "Адрес установки", "Камера"]
        return worksheet.get_all_records(expected_headers=expected_headers)

    def process_camera_data(self, data):
        """Обработка и группировка данных по адресам"""
        if not data:
            raise ValueError("В таблице нет данных!")

        address_data = defaultdict(lambda: defaultdict(int))
        all_models = set()
        object_codes = {}

        for row in data:
            code = row.get("Код объекта", "").strip()
            address = row.get("Адрес установки", "").strip()
            model = row.get("Камера", "").strip()
            if address and model:
                address_data[address][model] += 1
                all_models.add(model)
                object_codes[address] = code

        if not address_data:
            raise ValueError("Нет данных для формирования отчета!")

        return address_data, sorted(all_models), object_codes


class ExcelReportGenerator:
    """Класс для генерации Excel отчета"""

    def __init__(self, address_data, camera_models, object_codes, callback):
        self.address_data = address_data
        self.camera_models = camera_models
        self.object_codes = object_codes
        self.callback = callback

    def run(self):
        try:
            report = self.create_excel_report()
            filename = self.save_report(report)
            self.callback(filename, None)
        except Exception as e:
            self.callback(None, str(e))

    def create_excel_report(self):
        """Создание Excel файла с отчетом"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Лист1"

        # Стили оформления
        header_style = Font(bold=True)
        center_alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True)
        vertical_alignment = Alignment(
            textRotation=90, horizontal="center", vertical="center", wrap_text=True)
        border_style = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Выравнивание по центру для строк 1-3
        for row in range(1, 4):
            for col in range(1, 22):  # Предполагаем 21 столбец (A-X)
                ws.cell(row=row, column=col).alignment = center_alignment

        # Основной заголовок (A1:U1)
        ws.merge_cells('A1:U1')
        ws['A1'] = "Ведомость установленного и замонтированного оборудования по объекту:"
        ws['A1'].font = header_style

        # Название проекта (A2:U2)
        ws.merge_cells('A2:U2')
        ws['A2'] = "Реконструкция местных линий связи к объектам РСМОБ г. Бреста перекрестки, 8 этап"
        ws['A2'].font = header_style

        # Группы оборудования (строка 3)
        ws.merge_cells('C3:M3')
        c3 = ws['C3']
        c3.value = "Видеокамеры"
        c3.font = Font(bold=True)
        c3.alignment = center_alignment

        ws.merge_cells('Q3:S3')
        q3 = ws['Q3']
        q3.value = "Коммутаторы"
        q3.font = Font(bold=True)
        q3.alignment = center_alignment

        ws.merge_cells('T3:U3')
        t3 = ws['T3']
        t3.value = "Удлинитель"
        t3.font = Font(bold=True)
        t3.alignment = center_alignment

        # Заголовки столбцов (строка 4)
        headers = [
            "Код объекта",
            "АДРЕС",
            "(2 Мп) TIANDY TC-C32GS-I5EYCSD (2.8mm/V4.2)",
            "(2МР) IPC2122LB-ADF28KM-G",
            "DS-2CD1043G0-IUVSD 4mm",
            "DS-2CD2123G2-IUVSD 4mm",
            "DS-2CD2T23G2-2IUVSD",
            "DS-2CD2T23G2-2IUVSD 4mm",
            "DS-2CD3021G0-IUVSC 4mm",
            "DS-2CD3123G2-IUUVSC 6mm",
            "DS-2CD3626G2T-IZSUVSC (7-35mm)",
            "DS-2CD3626G2T-IZSUVSC 7-35mm",
            "DS-2CD3726G2T-IZSUVSC (7-35mm)",
            "DS-2DE5425IW-AEUVSC",
            "HIKVISION DS-2DE5425IW-A E (T5)",
            "Uniview IPC2122LE-ADF28KMC-WL",
            "Коммутатор ZTO L2S1900-4TP2S",
            "Коммутатор ZTO L2S 1900-8TP2S",
            "Коммутатор ZTO L2S 1900-16TP2S",
            "Инжектор питания (PoE) OPL-POE-Ex802 3at-100-IP67",
            "Удлинитель PoE ZTO POEEXT 100"
        ]

        # Заполняем строку 4 с разным выравниванием
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col, value=header)
            cell.font = header_style
            cell.border = border_style
            cell.alignment = vertical_alignment if col >= 3 else center_alignment

        ws.row_dimensions[4].height = 150

        # Заполнение данных
        row_idx = 5
        for address, counts in self.address_data.items():
            # Код объекта и адрес
            ws.cell(row=row_idx, column=1, value=self.object_codes.get(
                address, "")).alignment = center_alignment
            ws.cell(row=row_idx, column=2, value=address).alignment = Alignment(
                wrap_text=True)

            # Видеокамеры
            for model, quantity in counts.items():
                if model in headers:
                    col_idx = headers.index(model) + 1
                    ws.cell(row=row_idx, column=col_idx,
                            value=quantity).alignment = center_alignment

            # Коммутаторы и удлинители оставляем пустыми
            for col in range(17, 22):
                ws.cell(row=row_idx, column=col,
                        value="").alignment = center_alignment

            row_idx += 1

        # Итоговая строка
        total_row = row_idx
        ws.cell(row=total_row, column=1,
                value="ИТОГО:").alignment = center_alignment
        ws.cell(row=total_row, column=2, value="").alignment = center_alignment

        # Формулы суммирования
        for col in range(3, 22):
            col_letter = get_column_letter(col)
            ws.cell(row=total_row, column=col,
                    value=f"=SUM({col_letter}5:{col_letter}{total_row-1})").alignment = center_alignment

        # Подпись
        signature_row = total_row + 1
        ws.merge_cells(f'A{signature_row}:U{signature_row}')
        ws[f'A{signature_row}'] = "Подготовил: ведущий инженер ЛСС и АУ А.И. Козей"
        ws[f'A{signature_row}'].alignment = center_alignment

        # Форматирование границ для всех ячеек
        for row in ws.iter_rows():
            for cell in row:
                cell.border = border_style

        # Ширина столбцов
        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 50
        for col in range(3, 22):
            ws.column_dimensions[get_column_letter(col)].width = 8

        return wb

    def save_report(self, report):
        """Сохранение отчета в файл"""
        output_dir = Path("output")
        output_dir.mkdir(exist_ok=True)

        first_code = next(iter(self.object_codes.values()),
                          "") if self.object_codes else ""
        id_number = first_code[1] if len(
            first_code) > 1 and first_code[1].isdigit() else ""

        filename = f"Ведомость-{id_number}.xlsx" if id_number else f"Ведомость-{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        filepath = output_dir / filename
        report.save(filepath)
        return str(filepath)


class MainWindow(tk.Tk):
    """Главное окно приложения"""

    def __init__(self):
        super().__init__()
        self.title("Генератор ведомостей оборудования")
        self.geometry("800x600")

        # Инициализация переменных Tkinter
        self.progress_var = tk.DoubleVar()
        self.progress_var.set(0)

        # Загрузка конфигурации
        load_dotenv()
        self.spreadsheet_url = os.getenv("GOOGLE_SHEETS_URL", "")
        self.credentials_file = os.getenv(
            "CREDENTIALS_JSON", "credentials.json")
        self.sheet_name = os.getenv("SHEET_NAME", "Камеры")
        self.client_email = ""

        self.create_widgets()
        self.load_credentials_info()

    def create_widgets(self):
        """Создание элементов интерфейса"""
        main_frame = ttk.Frame(self)
        main_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)

        # Поля для настроек
        settings_frame = ttk.LabelFrame(main_frame, text="Настройки")
        settings_frame.pack(fill=tk.X, pady=5)

        # URL Google Sheets
        ttk.Label(settings_frame, text="URL Google Sheets:").grid(
            row=0, column=0, sticky=tk.W, padx=5, pady=2)
        self.url_entry = ttk.Entry(settings_frame, width=60)
        self.url_entry.insert(0, self.spreadsheet_url)
        self.url_entry.grid(row=0, column=1, sticky=tk.EW, padx=5, pady=2)

        # Имя листа
        ttk.Label(settings_frame, text="Имя листа:").grid(
            row=1, column=0, sticky=tk.W, padx=5, pady=2)
        self.sheet_entry = ttk.Entry(settings_frame)
        self.sheet_entry.insert(0, self.sheet_name)
        self.sheet_entry.grid(row=1, column=1, sticky=tk.EW, padx=5, pady=2)

        # Файл учетных данных
        ttk.Label(settings_frame, text="Файл учетных данных:").grid(
            row=2, column=0, sticky=tk.W, padx=5, pady=2)
        self.creds_entry = ttk.Entry(settings_frame)
        self.creds_entry.insert(0, self.credentials_file)
        self.creds_entry.grid(row=2, column=1, sticky=tk.EW, padx=5, pady=2)
        ttk.Button(settings_frame, text="Обзор...", command=self.browse_credentials_file).grid(
            row=2, column=2, padx=5, pady=2)

        # Client email - теперь в вертикальном расположении
        email_frame = ttk.Frame(settings_frame)
        email_frame.grid(row=3, column=0, columnspan=3,
                         sticky=tk.EW, padx=5, pady=2)

        ttk.Label(email_frame, text="Client Email:").pack(
            side=tk.LEFT, padx=(0, 5))
        self.email_entry = ttk.Entry(email_frame, state='readonly')
        self.email_entry.pack(side=tk.LEFT, expand=True,
                              fill=tk.X, padx=(0, 5))
        ttk.Button(email_frame, text="Копировать",
                   command=self.copy_client_email).pack(side=tk.LEFT)

        # Кнопки запуска - теперь в вертикальном расположении под email
        button_frame = ttk.Frame(settings_frame)
        button_frame.grid(row=4, column=0, columnspan=3,
                          sticky=tk.EW, padx=5, pady=(0, 5))

        self.run_button = ttk.Button(
            button_frame, text="Создать ведомость", command=self.run_report_generation)
        self.run_button.pack(fill=tk.X, pady=2)

        self.quick_run_button = ttk.Button(
            button_frame, text="Быстрый запуск", command=self.quick_run_report)
        self.quick_run_button.pack(fill=tk.X, pady=2)

        # Лог сообщений
        log_frame = ttk.LabelFrame(main_frame, text="Лог")
        log_frame.pack(fill=tk.BOTH, expand=True, pady=5)

        self.log_text = scrolledtext.ScrolledText(
            log_frame, wrap=tk.WORD, state='normal')
        self.log_text.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)

        # Прогресс бар
        self.progress = ttk.Progressbar(
            main_frame, variable=self.progress_var, maximum=100)
        self.progress.pack(fill=tk.X, pady=5)

        # Статус
        self.status_label = ttk.Label(main_frame, text="")
        self.status_label.pack()

        # Настройка веса строк и столбцов для правильного масштабирования
        settings_frame.columnconfigure(1, weight=1)
        main_frame.columnconfigure(0, weight=1)

    def load_credentials_info(self):
        """Загрузка информации из файла учетных данных"""
        try:
            if os.path.exists(self.credentials_file):
                with open(self.credentials_file, 'r') as f:
                    creds = json.load(f)
                    self.client_email = creds.get('client_email', '')
                    self.email_entry.config(state='normal')
                    self.email_entry.delete(0, tk.END)
                    self.email_entry.insert(0, self.client_email)
                    self.email_entry.config(state='readonly')
        except Exception as e:
            self.log_message(f"Ошибка загрузки файла учетных данных: {str(e)}")

    def browse_credentials_file(self):
        """Выбор файла учетных данных через диалог"""
        filename = filedialog.askopenfilename(
            filetypes=[("JSON Files", "*.json")])
        if filename:
            self.creds_entry.delete(0, tk.END)
            self.creds_entry.insert(0, filename)
            self.credentials_file = filename
            self.load_credentials_info()

    def copy_client_email(self):
        """Копирование client_email в буфер обмена"""
        self.clipboard_clear()
        self.clipboard_append(self.client_email)
        self.log_message("Client email скопирован в буфер обмена")

    def log_message(self, message):
        """Добавление сообщения в лог"""
        self.log_text.config(state='normal')
        self.log_text.insert(tk.END, message + "\n")
        self.log_text.config(state='disabled')
        self.log_text.see(tk.END)

    def update_progress(self, value):
        """Обновление прогресс бара"""
        self.progress_var.set(value)
        self.update_idletasks()

    def run_report_generation(self):
        """Запуск процесса генерации отчета"""
        self.spreadsheet_url = self.url_entry.get().strip()
        self.sheet_name = self.sheet_entry.get().strip()
        self.credentials_file = self.creds_entry.get().strip()

        if not all([self.spreadsheet_url, self.credentials_file]):
            self.log_message("Ошибка: Не заданы все необходимые параметры!")
            return

        self.log_message("Начало обработки данных...")
        self.progress_var.set(0)
        self.run_button.config(state='disabled')
        self.quick_run_button.config(state='disabled')

        # Запускаем в отдельном потоке
        worker = GoogleSheetsWorker(
            self.spreadsheet_url,
            self.credentials_file,
            self.sheet_name,
            self.on_data_processed
        )

        thread = threading.Thread(target=worker.run)
        thread.daemon = True
        thread.start()

        # Анимация прогресс-бара
        self.animate_progress()

    def quick_run_report(self):
        """Быстрый запуск с текущими параметрами"""
        self.log_message("Быстрый запуск создания отчета...")
        self.run_report_generation()

    def animate_progress(self):
        """Анимация прогресс-бара"""
        current = self.progress_var.get()
        if current < 100:
            self.progress_var.set(current + 1)
            self.after(100, self.animate_progress)

    def on_data_processed(self, address_data, camera_models, object_codes, error):
        """Обработка завершения получения данных"""
        if error:
            self.on_error(error)
            return

        self.log_message(f"Обработано {len(address_data)} адресов")
        self.log_message(f"Найдено {len(camera_models)} моделей камер")

        # Запускаем генерацию отчета
        report_worker = ExcelReportGenerator(
            address_data,
            camera_models,
            object_codes,
            self.on_report_generated
        )

        thread = threading.Thread(target=report_worker.run)
        thread.daemon = True
        thread.start()

    def on_report_generated(self, filename, error):
        """Обработка завершения генерации отчета"""
        if error:
            self.on_error(error)
            return

        self.run_button.config(state='normal')
        self.quick_run_button.config(state='normal')
        self.status_label.config(text=f"Отчет сохранен: {filename}")
        self.log_message(f"Отчет успешно создан: {filename}")

        messagebox.showinfo(
            "Успех", f"Отчет успешно создан!\nФайл сохранен как: {filename}")

    def on_error(self, error_message):
        """Обработка ошибок"""
        self.log_message(error_message)
        self.run_button.config(state='normal')
        self.quick_run_button.config(state='normal')
        self.progress_var.set(0)

        messagebox.showerror("Ошибка", error_message)


def main():
    """Точка входа в приложение"""
    # Создаем папку для отчетов, если ее нет
    output_dir = Path("output")
    output_dir.mkdir(exist_ok=True)

    app = MainWindow()
    app.mainloop()


if __name__ == "__main__":
    main()
